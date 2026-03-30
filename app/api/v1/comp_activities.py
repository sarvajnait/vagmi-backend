from datetime import UTC, datetime
from io import BytesIO
import re
from typing import Optional
import xml.etree.ElementTree as ET
import zipfile

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, Query, UploadFile
from pydantic import BaseModel
from sqlalchemy import case, func
from sqlmodel import select
from sqlmodel.ext.asyncio.session import AsyncSession

from app.api.v1.auth import get_current_user
from app.api.v1.admin.auth import get_current_user as get_current_admin
from app.models import ActivityGenerationJob
from app.models.comp_activities import (
    CompActivityGroup, CompActivityGroupCreate, CompChapterActivity, CompTopic,
    CompActivityPlaySession, CompActivityAnswer,
)
from app.models.user import User
from app.models.admin import Admin
from app.services.activity_jobs import enqueue_activity_job
from app.services.database import get_session
from app.utils.files import upload_to_do, delete_from_do

from openpyxl import load_workbook

router = APIRouter()


class OrderUpdate(BaseModel):
    ids: list[int]


class CompActivityGroupUpdate(BaseModel):
    name: Optional[str] = None
    timer_seconds: Optional[int] = None
    sort_order: Optional[int] = None


class GenerateCompActivitiesRequest(BaseModel):
    comp_chapter_id: Optional[int] = None
    sub_chapter_id: Optional[int] = None
    topic_titles: list[str] = []
    mcq_count: int = 5
    descriptive_count: int = 5
    activity_group_id: Optional[int] = None


class CompTopicsRequest(BaseModel):
    comp_chapter_id: Optional[int] = None
    sub_chapter_id: Optional[int] = None


class PublishRequest(BaseModel):
    ids: list[int]
    is_published: bool = True


def sort_ordering(model):
    return [
        case((model.sort_order == None, 1), else_=0),
        model.sort_order,
        model.created_at,
    ]


def _resolve_fk(comp_chapter_id: Optional[int], sub_chapter_id: Optional[int]):
    if comp_chapter_id is None and sub_chapter_id is None:
        raise HTTPException(status_code=400, detail="Either comp_chapter_id or sub_chapter_id must be provided")
    return comp_chapter_id, sub_chapter_id


_HEADER_ALIASES = {
    "topic": "topic_name",
    "topic_name": "topic_name",
    "topicname": "topic_name",
    "question": "question_text",
    "question_text": "question_text",
    "questiontext": "question_text",
    "option_a": "option_a",
    "option_a_text": "option_a",
    "option1": "option_a",
    "option_1": "option_a",
    "a": "option_a",
    "option_b": "option_b",
    "option_b_text": "option_b",
    "option2": "option_b",
    "option_2": "option_b",
    "b": "option_b",
    "option_c": "option_c",
    "option_c_text": "option_c",
    "option3": "option_c",
    "option_3": "option_c",
    "c": "option_c",
    "option_d": "option_d",
    "option_d_text": "option_d",
    "option4": "option_d",
    "option_4": "option_d",
    "d": "option_d",
    "correct_option_index": "correct_option_index",
    "correct_index": "correct_option_index",
    "correct_option": "correct_option_index",
    "correct_answer": "correct_option_index",
    "correct_option_letter": "correct_option_letter",
    "correct_letter": "correct_option_letter",
    "answer_description": "answer_description",
    "explanation": "answer_description",
    "answer_explanation": "answer_description",
    "is_published": "is_published",
    "published": "is_published",
}


def _normalize_header(value: str) -> str:
    key = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")
    return _HEADER_ALIASES.get(key, key)


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _parse_bool(value) -> Optional[bool]:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "yes", "1", "y"}:
        return True
    if text in {"false", "no", "0", "n"}:
        return False
    return None


def _parse_correct_index(value) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, int):
        return value if value in {1, 2, 3, 4} else None
    text = str(value).strip().lower()
    if text in {"a", "(a)"}:
        return 1
    if text in {"b", "(b)"}:
        return 2
    if text in {"c", "(c)"}:
        return 3
    if text in {"d", "(d)"}:
        return 4
    try:
        num = int(float(text))
        return num if num in {1, 2, 3, 4} else None
    except Exception:
        return None


def _extract_docx_paragraphs(content: bytes) -> list[str]:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with zipfile.ZipFile(BytesIO(content), "r") as archive:
        xml_data = archive.read("word/document.xml")
    root = ET.fromstring(xml_data)
    paragraphs = []
    for para in root.findall(".//w:p", ns):
        text = "".join(node.text or "" for node in para.findall(".//w:t", ns)).strip()
        if text:
            paragraphs.append(text)
    return paragraphs


def _parse_option_line(option_line: str, expected_letters: tuple[str, str]) -> dict[str, str]:
    first, second = expected_letters
    first_match = re.search(rf"\({first}\)\s*", option_line, flags=re.IGNORECASE)
    second_match = re.search(rf"\({second}\)\s*", option_line, flags=re.IGNORECASE)
    if not first_match or not second_match or second_match.start() <= first_match.end():
        return {}
    return {
        f"option_{first}": option_line[first_match.end():second_match.start()].strip(),
        f"option_{second}": option_line[second_match.end():].strip(),
    }


def _parse_docx_rows(content: bytes) -> list[dict]:
    paragraphs = _extract_docx_paragraphs(content)
    topic_re = re.compile(r"^TOPIC\s+\d+\s*:\s*(.+)$", re.IGNORECASE)
    number_re = re.compile(r"^\d+$")
    answer_re = re.compile(r"^\(([a-d])\)$", re.IGNORECASE)
    terminal_markers = {
        "SELF-ASSESSMENT SCORE TRACKER",
        "FINAL REVISION STRATEGY",
        "HOW TO REVISE THIS MCQ BANK IN 7 DAYS",
    }

    rows: list[dict] = []
    i = 0
    while i < len(paragraphs):
        topic_match = topic_re.match(paragraphs[i])
        if not topic_match:
            i += 1
            continue

        topic_name = topic_match.group(1).strip()
        i += 1

        while i < len(paragraphs) and paragraphs[i] != "Questions":
            if topic_re.match(paragraphs[i]):
                break
            i += 1
        if i >= len(paragraphs) or topic_re.match(paragraphs[i]):
            continue

        i += 3
        question_map: dict[int, dict] = {}

        while i < len(paragraphs) and paragraphs[i] != "Answer Key with Explanations":
            if topic_re.match(paragraphs[i]):
                break
            if not number_re.match(paragraphs[i]):
                i += 1
                continue

            qno = int(paragraphs[i])
            if i + 3 >= len(paragraphs):
                raise ValueError(f"Incomplete question block for topic '{topic_name}' question {qno}")

            question_text = paragraphs[i + 1].strip()
            options = {}
            options.update(_parse_option_line(paragraphs[i + 2], ("a", "b")))
            options.update(_parse_option_line(paragraphs[i + 3], ("c", "d")))

            if len(options) != 4:
                raise ValueError(f"Could not parse 4 options for topic '{topic_name}' question {qno}")

            question_map[qno] = {
                "topic_name": topic_name,
                "question_text": question_text,
                "option_a": options.get("option_a", ""),
                "option_b": options.get("option_b", ""),
                "option_c": options.get("option_c", ""),
                "option_d": options.get("option_d", ""),
                "source_row": f"{topic_name} Q{qno}",
            }
            i += 4

        if i >= len(paragraphs) or paragraphs[i] != "Answer Key with Explanations":
            raise ValueError(f"Answer key not found for topic '{topic_name}'")

        i += 3
        while i < len(paragraphs):
            if paragraphs[i] in terminal_markers:
                break
            if topic_re.match(paragraphs[i]):
                break
            if not number_re.match(paragraphs[i]):
                i += 1
                continue

            qno = int(paragraphs[i])
            if i + 2 >= len(paragraphs):
                raise ValueError(f"Incomplete answer block for topic '{topic_name}' question {qno}")
            answer_match = answer_re.match(paragraphs[i + 1].strip())
            if not answer_match:
                raise ValueError(f"Invalid answer format for topic '{topic_name}' question {qno}")
            explanation = paragraphs[i + 2].strip()
            if qno not in question_map:
                raise ValueError(f"Answer exists without question for topic '{topic_name}' question {qno}")
            question_map[qno]["correct_option_index"] = _parse_correct_index(answer_match.group(1))
            question_map[qno]["answer_description"] = explanation
            i += 3

        rows.extend(question_map[qno] for qno in sorted(question_map))

    if not rows:
        raise ValueError("No MCQ content could be parsed from the DOCX file")
    return rows


def _parse_xlsx_rows(content: bytes) -> list[dict]:
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active

    headers = {}
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError("Empty Excel file")
    for idx, raw in enumerate(header_row):
        if raw is None:
            continue
        headers[_normalize_header(raw)] = idx

    required = {"question_text", "option_a", "option_b", "option_c", "option_d"}
    missing = [header for header in required if header not in headers]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all((cell is None or str(cell).strip() == "") for cell in row):
            continue
        rows.append({
            "topic_name": _cell_str(row[headers["topic_name"]]) if "topic_name" in headers else "",
            "question_text": _cell_str(row[headers["question_text"]]) if "question_text" in headers else "",
            "option_a": _cell_str(row[headers["option_a"]]) if "option_a" in headers else "",
            "option_b": _cell_str(row[headers["option_b"]]) if "option_b" in headers else "",
            "option_c": _cell_str(row[headers["option_c"]]) if "option_c" in headers else "",
            "option_d": _cell_str(row[headers["option_d"]]) if "option_d" in headers else "",
            "correct_option_index": (
                _parse_correct_index(row[headers["correct_option_index"]])
                if "correct_option_index" in headers
                else None
            ),
            "correct_option_letter": (
                _parse_correct_index(row[headers["correct_option_letter"]])
                if "correct_option_letter" in headers
                else None
            ),
            "answer_description": _cell_str(row[headers["answer_description"]]) if "answer_description" in headers else "",
            "is_published": _parse_bool(row[headers["is_published"]]) if "is_published" in headers else None,
            "source_row": row_idx,
        })
    return rows


# ============================================================
# Activity Groups
# ============================================================

@router.post("/activity-groups")
async def create_comp_activity_group(
    payload: CompActivityGroupCreate,
    session: AsyncSession = Depends(get_session),
):
    try:
        comp_chapter_id, sub_chapter_id = _resolve_fk(payload.comp_chapter_id, payload.sub_chapter_id)
        if not payload.name or not payload.name.strip():
            raise HTTPException(status_code=400, detail="Group name is required")
        filter_col = CompActivityGroup.comp_chapter_id if comp_chapter_id else CompActivityGroup.sub_chapter_id
        filter_val = comp_chapter_id or sub_chapter_id
        if payload.sort_order is None:
            _result = await session.exec(select(func.max(CompActivityGroup.sort_order)).where(filter_col == filter_val))
            max_order = _result.first()
            if isinstance(max_order, tuple):
                max_order = max_order[0]
            sort_order = (max_order or 0) + 1
        else:
            sort_order = payload.sort_order
        group = CompActivityGroup(
            name=payload.name.strip(),
            comp_chapter_id=comp_chapter_id,
            sub_chapter_id=sub_chapter_id,
            timer_seconds=payload.timer_seconds,
            sort_order=sort_order,
        )
        session.add(group)
        await session.commit()
        await session.refresh(group)
        return {"message": "Activity group created", "data": group.dict()}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/activity-groups")
async def get_comp_activity_groups(
    comp_chapter_id: Optional[int] = None,
    sub_chapter_id: Optional[int] = None,
    session: AsyncSession = Depends(get_session),
):
    try:
        query = select(CompActivityGroup)
        if comp_chapter_id is not None:
            query = query.where(CompActivityGroup.comp_chapter_id == comp_chapter_id)
        elif sub_chapter_id is not None:
            query = query.where(CompActivityGroup.sub_chapter_id == sub_chapter_id)
        query = query.order_by(*sort_ordering(CompActivityGroup))
        result = await session.exec(query)
        groups = result.all()
        groups_data = []
        for group in groups:
            _count = await session.exec(select(func.count()).where(CompChapterActivity.activity_group_id == group.id))
            activity_count = _count.first()
            if isinstance(activity_count, tuple):
                activity_count = activity_count[0]
            groups_data.append({**group.dict(), "activity_count": activity_count or 0})
        return {"data": groups_data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/activity-groups/{group_id}")
async def get_comp_activity_group(group_id: int, session: AsyncSession = Depends(get_session)):
    group = await session.get(CompActivityGroup, group_id)
    if not group:
        raise HTTPException(status_code=404, detail="Activity group not found")
    _count = await session.exec(select(func.count()).where(CompChapterActivity.activity_group_id == group_id))
    activity_count = _count.first()
    if isinstance(activity_count, tuple):
        activity_count = activity_count[0]
    return {"data": {**group.dict(), "activity_count": activity_count or 0}}


@router.put("/activity-groups/{group_id}")
async def update_comp_activity_group(
    group_id: int,
    payload: CompActivityGroupUpdate,
    session: AsyncSession = Depends(get_session),
):
    try:
        group = await session.get(CompActivityGroup, group_id)
        if not group:
            raise HTTPException(status_code=404, detail="Activity group not found")
        if payload.name is not None:
            if not payload.name.strip():
                raise HTTPException(status_code=400, detail="Group name cannot be empty")
            group.name = payload.name.strip()
        if payload.timer_seconds is not None:
            group.timer_seconds = payload.timer_seconds
        if payload.sort_order is not None:
            group.sort_order = payload.sort_order
        session.add(group)
        await session.commit()
        await session.refresh(group)
        return {"message": "Activity group updated", "data": group.dict()}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/activity-groups/{group_id}")
async def delete_comp_activity_group(group_id: int, session: AsyncSession = Depends(get_session)):
    try:
        group = await session.get(CompActivityGroup, group_id)
        if not group:
            raise HTTPException(status_code=404, detail="Activity group not found")
        await session.delete(group)
        await session.commit()
        return {"message": "Activity group deleted"}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/activity-groups/order")
async def reorder_comp_activity_groups(
    payload: OrderUpdate,
    comp_chapter_id: Optional[int] = Query(None),
    sub_chapter_id: Optional[int] = Query(None),
    session: AsyncSession = Depends(get_session),
):
    filter_col = CompActivityGroup.comp_chapter_id if comp_chapter_id else CompActivityGroup.sub_chapter_id
    filter_val = comp_chapter_id or sub_chapter_id
    _result = await session.exec(select(CompActivityGroup).where(filter_col == filter_val, CompActivityGroup.id.in_(payload.ids)))
    groups = _result.all()
    if len(groups) != len(payload.ids):
        raise HTTPException(status_code=400, detail="Invalid activity group ids")
    group_map = {g.id: g for g in groups}
    for index, gid in enumerate(payload.ids, start=1):
        group_map[gid].sort_order = index
        session.add(group_map[gid])
    await session.commit()
    return {"message": "Activity group order updated"}


# ============================================================
# Activities
# ============================================================

@router.post("/activities")
async def create_comp_activity(
    activity_group_id: int = Form(...),
    comp_chapter_id: Optional[int] = Form(None),
    sub_chapter_id: Optional[int] = Form(None),
    type: str = Form(...),
    question_text: str = Form(...),
    options: Optional[list[str]] = Form(None),
    correct_option_index: Optional[int] = Form(None),
    answer_text: Optional[str] = Form(None),
    answer_description: Optional[str] = Form(None),
    is_published: bool = Form(True),
    sort_order: Optional[int] = Form(None),
    answer_image: UploadFile | None = File(None),
    session: AsyncSession = Depends(get_session),
):
    try:
        group = await session.get(CompActivityGroup, activity_group_id)
        if not group:
            raise HTTPException(status_code=404, detail="Activity group not found")

        cleaned_options = [opt.strip() for opt in options] if options else None

        if sort_order is None:
            _result = await session.exec(
                select(func.max(CompChapterActivity.sort_order)).where(
                    CompChapterActivity.activity_group_id == activity_group_id
                )
            )
            max_order = _result.first()
            if isinstance(max_order, tuple):
                max_order = max_order[0]
            sort_order = (max_order or 0) + 1

        answer_image_url = None
        if answer_image:
            folder_id = comp_chapter_id or sub_chapter_id
            do_path = f"comp/chapters/{folder_id}/activities/answers"
            answer_image_url = upload_to_do(answer_image, do_path)

        activity = CompChapterActivity(
            activity_group_id=activity_group_id,
            comp_chapter_id=comp_chapter_id,
            sub_chapter_id=sub_chapter_id,
            type=type,
            question_text=question_text.strip(),
            options=cleaned_options if type == "mcq" else None,
            correct_option_index=correct_option_index if type == "mcq" else None,
            answer_text=answer_text.strip() if answer_text and type == "descriptive" else None,
            answer_description=answer_description.strip() if answer_description else None,
            answer_image_url=answer_image_url,
            is_published=is_published,
            sort_order=sort_order,
        )
        session.add(activity)
        await session.commit()
        await session.refresh(activity)
        return {"message": "Activity created", "data": activity.dict()}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/activities")
async def get_comp_activities(
    activity_group_id: Optional[int] = None,
    comp_chapter_id: Optional[int] = None,
    sub_chapter_id: Optional[int] = None,
    status: str = Query("all"),
    session: AsyncSession = Depends(get_session),
):
    try:
        query = select(CompChapterActivity)
        if activity_group_id is not None:
            query = query.where(CompChapterActivity.activity_group_id == activity_group_id)
        if comp_chapter_id is not None:
            query = query.where(CompChapterActivity.comp_chapter_id == comp_chapter_id)
        elif sub_chapter_id is not None:
            query = query.where(CompChapterActivity.sub_chapter_id == sub_chapter_id)
        if status == "published":
            query = query.where(CompChapterActivity.is_published == True)
        elif status == "draft":
            query = query.where(CompChapterActivity.is_published == False)
        query = query.order_by(*sort_ordering(CompChapterActivity))
        result = await session.exec(query)
        return {"data": [a.dict() for a in result.all()]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/activities/{activity_id}")
async def update_comp_activity(
    activity_id: int,
    activity_group_id: Optional[int] = Form(None),
    type: Optional[str] = Form(None),
    question_text: Optional[str] = Form(None),
    options: Optional[list[str]] = Form(None),
    correct_option_index: Optional[int] = Form(None),
    answer_text: Optional[str] = Form(None),
    answer_description: Optional[str] = Form(None),
    is_published: Optional[bool] = Form(None),
    sort_order: Optional[int] = Form(None),
    answer_image: UploadFile | None = File(None),
    session: AsyncSession = Depends(get_session),
):
    try:
        activity = await session.get(CompChapterActivity, activity_id)
        if not activity:
            raise HTTPException(status_code=404, detail="Activity not found")
        if activity_group_id is not None:
            activity.activity_group_id = activity_group_id
        if type is not None:
            activity.type = type
        if question_text is not None:
            activity.question_text = question_text.strip()
        if options is not None:
            activity.options = [opt.strip() for opt in options]
        if correct_option_index is not None:
            activity.correct_option_index = correct_option_index
        if answer_text is not None:
            activity.answer_text = answer_text.strip()
        if answer_description is not None:
            activity.answer_description = answer_description.strip() if answer_description else None
        if is_published is not None:
            activity.is_published = is_published
        if sort_order is not None:
            activity.sort_order = sort_order
        if answer_image:
            folder_id = activity.comp_chapter_id or activity.sub_chapter_id
            new_url = upload_to_do(answer_image, f"comp/chapters/{folder_id}/activities/answers")
            if activity.answer_image_url:
                delete_from_do(activity.answer_image_url)
            activity.answer_image_url = new_url
        session.add(activity)
        await session.commit()
        await session.refresh(activity)
        return {"message": "Activity updated", "data": activity.dict()}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/activities/{activity_id}")
async def delete_comp_activity(activity_id: int, session: AsyncSession = Depends(get_session)):
    try:
        activity = await session.get(CompChapterActivity, activity_id)
        if not activity:
            raise HTTPException(status_code=404, detail="Activity not found")
        if activity.answer_image_url:
            delete_from_do(activity.answer_image_url)
        await session.delete(activity)
        await session.commit()
        return {"message": "Activity deleted"}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/activities/import")
async def import_comp_activities(
    file: UploadFile = File(...),
    import_mode: str = Form("use_group"),
    activity_group_id: Optional[int] = Form(None),
    comp_chapter_id: Optional[int] = Form(None),
    sub_chapter_id: Optional[int] = Form(None),
    session: AsyncSession = Depends(get_session),
):
    filename = (file.filename or "").lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".docx")):
        raise HTTPException(status_code=400, detail="Only .xlsx and .docx files are supported")

    if import_mode not in {"use_group", "create_groups"}:
        raise HTTPException(status_code=400, detail="Invalid import_mode")

    group = None
    if import_mode == "use_group":
        if not activity_group_id:
            raise HTTPException(status_code=400, detail="activity_group_id is required for use_group")
        group = await session.get(CompActivityGroup, activity_group_id)
        if not group:
            raise HTTPException(status_code=404, detail="Activity group not found")
        if comp_chapter_id is None and sub_chapter_id is None:
            comp_chapter_id = group.comp_chapter_id
            sub_chapter_id = group.sub_chapter_id
    else:
        comp_chapter_id, sub_chapter_id = _resolve_fk(comp_chapter_id, sub_chapter_id)

    try:
        content = await file.read()
        if filename.endswith(".xlsx"):
            parsed_rows = _parse_xlsx_rows(content)
        else:
            parsed_rows = _parse_docx_rows(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read import file: {e}")

    created = 0
    total_rows = len(parsed_rows)
    errors = []
    groups_created = set()
    groups_used = set()
    group_cache = {}
    group_sort_cache = {}

    async def _get_next_sort(gid: int) -> int:
        if gid in group_sort_cache:
            group_sort_cache[gid] += 1
            return group_sort_cache[gid]
        _result = await session.exec(
            select(func.max(CompChapterActivity.sort_order)).where(CompChapterActivity.activity_group_id == gid)
        )
        max_order = _result.first()
        if isinstance(max_order, tuple):
            max_order = max_order[0]
        group_sort_cache[gid] = (max_order or 0) + 1
        return group_sort_cache[gid]

    async def _get_or_create_group(topic_name: str) -> CompActivityGroup:
        key = topic_name.strip().lower()
        if key in group_cache:
            return group_cache[key]
        filter_col = CompActivityGroup.comp_chapter_id if comp_chapter_id else CompActivityGroup.sub_chapter_id
        filter_val = comp_chapter_id or sub_chapter_id
        _result = await session.exec(
            select(CompActivityGroup).where(
                filter_col == filter_val,
                func.lower(CompActivityGroup.name) == key,
            )
        )
        existing = _result.first()
        if existing:
            group_cache[key] = existing
            return existing
        _max = await session.exec(select(func.max(CompActivityGroup.sort_order)).where(filter_col == filter_val))
        max_order = _max.first()
        if isinstance(max_order, tuple):
            max_order = max_order[0]
        new_group = CompActivityGroup(
            name=topic_name.strip(),
            comp_chapter_id=comp_chapter_id,
            sub_chapter_id=sub_chapter_id,
            sort_order=(max_order or 0) + 1,
        )
        session.add(new_group)
        await session.commit()
        await session.refresh(new_group)
        group_cache[key] = new_group
        groups_created.add(new_group.id)
        return new_group

    for row in parsed_rows:
        row_ref = row.get("source_row", "?")
        try:
            question_text = _cell_str(row.get("question_text"))
            option_a = _cell_str(row.get("option_a"))
            option_b = _cell_str(row.get("option_b"))
            option_c = _cell_str(row.get("option_c"))
            option_d = _cell_str(row.get("option_d"))
            answer_description = _cell_str(row.get("answer_description"))
            topic_name = _cell_str(row.get("topic_name"))
            correct_index = row.get("correct_option_index")
            if correct_index is None:
                correct_index = row.get("correct_option_letter")

            if not question_text:
                raise ValueError("question_text is required")
            options = [option_a, option_b, option_c, option_d]
            if any(not opt for opt in options):
                raise ValueError("All 4 options are required")
            if correct_index not in {1, 2, 3, 4}:
                raise ValueError("correct_option_index must be 1-4 or a/b/c/d")

            if import_mode == "create_groups":
                if not topic_name:
                    raise ValueError("topic_name is required when creating groups")
                target_group = await _get_or_create_group(topic_name)
            else:
                target_group = group

            if target_group.comp_chapter_id and comp_chapter_id and target_group.comp_chapter_id != comp_chapter_id:
                raise ValueError("Activity group does not belong to the selected chapter")
            if target_group.sub_chapter_id and sub_chapter_id and target_group.sub_chapter_id != sub_chapter_id:
                raise ValueError("Activity group does not belong to the selected sub-chapter")

            is_published = row.get("is_published")
            sort_order = await _get_next_sort(target_group.id)

            activity = CompChapterActivity(
                activity_group_id=target_group.id,
                comp_chapter_id=comp_chapter_id or target_group.comp_chapter_id,
                sub_chapter_id=sub_chapter_id or target_group.sub_chapter_id,
                type="mcq",
                question_text=question_text,
                options=options,
                correct_option_index=int(correct_index),
                answer_text=None,
                answer_description=answer_description or None,
                is_published=True if is_published is None else is_published,
                sort_order=sort_order,
            )
            session.add(activity)
            groups_used.add(target_group.id)
            created += 1
        except Exception as e:
            errors.append({"row": row_ref, "error": str(e)})

    try:
        await session.commit()
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to import activities: {e}")

    return {
        "message": "Import completed",
        "data": {
            "total_rows": total_rows,
            "created": created,
            "errors": errors,
            "groups_used": len(groups_used),
            "groups_created": len(groups_created),
        },
    }


@router.put("/activities/order")
async def reorder_comp_activities(
    payload: OrderUpdate,
    activity_group_id: int = Query(...),
    session: AsyncSession = Depends(get_session),
):
    _result = await session.exec(
        select(CompChapterActivity).where(
            CompChapterActivity.activity_group_id == activity_group_id,
            CompChapterActivity.id.in_(payload.ids),
        )
    )
    activities = _result.all()
    if len(activities) != len(payload.ids):
        raise HTTPException(status_code=400, detail="Invalid activity ids for group")
    a_map = {a.id: a for a in activities}
    for index, aid in enumerate(payload.ids, start=1):
        a_map[aid].sort_order = index
        session.add(a_map[aid])
    await session.commit()
    return {"message": "Activity order updated"}


# ============================================================
# Topics
# ============================================================

@router.get("/topics")
async def get_comp_topics(
    comp_chapter_id: Optional[int] = None,
    sub_chapter_id: Optional[int] = None,
    session: AsyncSession = Depends(get_session),
):
    query = select(CompTopic)
    if comp_chapter_id is not None:
        query = query.where(CompTopic.comp_chapter_id == comp_chapter_id)
    elif sub_chapter_id is not None:
        query = query.where(CompTopic.sub_chapter_id == sub_chapter_id)
    query = query.order_by(*sort_ordering(CompTopic))
    result = await session.exec(query)
    return {"data": [t.dict() for t in result.all()]}


class CompTopicCreate(BaseModel):
    title: str
    comp_chapter_id: Optional[int] = None
    sub_chapter_id: Optional[int] = None


@router.post("/topics")
async def create_comp_topic(payload: CompTopicCreate, session: AsyncSession = Depends(get_session)):
    comp_chapter_id, sub_chapter_id = _resolve_fk(payload.comp_chapter_id, payload.sub_chapter_id)
    topic = CompTopic(title=payload.title.strip(), comp_chapter_id=comp_chapter_id, sub_chapter_id=sub_chapter_id)
    session.add(topic)
    await session.commit()
    await session.refresh(topic)
    return {"data": topic.dict()}


@router.delete("/topics/{topic_id}")
async def delete_comp_topic(topic_id: int, session: AsyncSession = Depends(get_session)):
    topic = await session.get(CompTopic, topic_id)
    if not topic:
        raise HTTPException(status_code=404, detail="Topic not found")
    await session.delete(topic)
    await session.commit()
    return {"data": {"deleted": True}}


@router.post("/topics/ai/generate")
async def ai_generate_comp_topics(
    payload: CompTopicsRequest,
    background_tasks: BackgroundTasks,
    session: AsyncSession = Depends(get_session),
):
    comp_chapter_id, sub_chapter_id = _resolve_fk(payload.comp_chapter_id, payload.sub_chapter_id)
    job = ActivityGenerationJob(
        job_type="comp_topics",
        status="pending",
        payload={"comp_chapter_id": comp_chapter_id, "sub_chapter_id": sub_chapter_id},
    )
    session.add(job)
    await session.commit()
    await session.refresh(job)
    background_tasks.add_task(enqueue_activity_job, job.id)
    return {"data": {"job_id": job.id, "status": job.status}}


@router.post("/topics/ai/save")
async def ai_save_comp_topics(
    payload: CompTopicsRequest,
    background_tasks: BackgroundTasks,
    session: AsyncSession = Depends(get_session),
):
    comp_chapter_id, sub_chapter_id = _resolve_fk(payload.comp_chapter_id, payload.sub_chapter_id)
    job = ActivityGenerationJob(
        job_type="comp_topics_save",
        status="pending",
        payload={"comp_chapter_id": comp_chapter_id, "sub_chapter_id": sub_chapter_id},
    )
    session.add(job)
    await session.commit()
    await session.refresh(job)
    background_tasks.add_task(enqueue_activity_job, job.id)
    return {"data": {"job_id": job.id, "status": job.status}}


@router.post("/activities/ai/generate")
async def ai_generate_comp_activities(
    payload: GenerateCompActivitiesRequest,
    background_tasks: BackgroundTasks,
    session: AsyncSession = Depends(get_session),
):
    comp_chapter_id, sub_chapter_id = _resolve_fk(payload.comp_chapter_id, payload.sub_chapter_id)
    mcq_count = min(max(payload.mcq_count, 0), 20)
    descriptive_count = min(max(payload.descriptive_count, 0), 20)
    if mcq_count + descriptive_count == 0:
        raise HTTPException(status_code=400, detail="Counts cannot both be zero")
    job = ActivityGenerationJob(
        job_type="comp_activities",
        status="pending",
        payload={
            "comp_chapter_id": comp_chapter_id,
            "sub_chapter_id": sub_chapter_id,
            "topic_titles": payload.topic_titles,
            "mcq_count": mcq_count,
            "descriptive_count": descriptive_count,
            "activity_group_id": payload.activity_group_id,
        },
    )
    session.add(job)
    await session.commit()
    await session.refresh(job)
    background_tasks.add_task(enqueue_activity_job, job.id)
    return {"data": {"job_id": job.id, "status": job.status}}


@router.get("/ai/jobs/{job_id}")
async def get_comp_ai_job(job_id: int, session: AsyncSession = Depends(get_session)):
    job = await session.get(ActivityGenerationJob, job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return {"data": job.dict()}


# ============================================================
# Publish
# ============================================================

@router.post("/activities/publish")
async def publish_comp_activities(
    payload: PublishRequest,
    _: Admin = Depends(get_current_admin),
    session: AsyncSession = Depends(get_session),
):
    if not payload.ids:
        raise HTTPException(status_code=400, detail="No ids provided")
    _result = await session.exec(
        select(CompChapterActivity).where(CompChapterActivity.id.in_(payload.ids))
    )
    activities = _result.all()
    if len(activities) != len(payload.ids):
        raise HTTPException(status_code=400, detail="Invalid activity ids")
    for activity in activities:
        activity.is_published = payload.is_published
        session.add(activity)
    await session.commit()
    return {"message": "Activities updated"}


# ============================================================
# Progress & Sessions (user-facing)
# ============================================================

class CompSessionCreate(BaseModel):
    comp_chapter_id: Optional[int] = None
    sub_chapter_id: Optional[int] = None


class CompAnswerSubmit(BaseModel):
    activity_id: int
    selected_option_index: Optional[int] = None


@router.get("/progress")
async def get_comp_activity_progress(
    comp_chapter_id: Optional[int] = Query(None),
    sub_chapter_id: Optional[int] = Query(None),
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    filter_col = CompChapterActivity.comp_chapter_id if comp_chapter_id else CompChapterActivity.sub_chapter_id
    filter_val = comp_chapter_id or sub_chapter_id
    _result = await session.exec(
        select(CompChapterActivity)
        .where(filter_col == filter_val, CompChapterActivity.is_published == True)
        .order_by(*sort_ordering(CompChapterActivity))
    )
    activities = _result.all()
    if not activities:
        return {"data": []}

    activity_ids = [a.id for a in activities]
    answered_result = await session.exec(
        select(CompActivityAnswer.activity_id)
        .join(CompActivityPlaySession)
        .where(
            CompActivityPlaySession.user_id == current_user.id,
            CompActivityAnswer.activity_id.in_(activity_ids),
        )
        .distinct()
    )
    answered_ids = {row[0] if isinstance(row, tuple) else row for row in answered_result.all()}
    return {"data": [{**a.dict(), "completed": a.id in answered_ids} for a in activities]}


@router.post("/sessions")
async def create_comp_session(
    payload: CompSessionCreate,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    comp_chapter_id, sub_chapter_id = _resolve_fk(payload.comp_chapter_id, payload.sub_chapter_id)
    filter_col = CompChapterActivity.comp_chapter_id if comp_chapter_id else CompChapterActivity.sub_chapter_id
    filter_val = comp_chapter_id or sub_chapter_id

    _count_result = await session.exec(
        select(func.count()).where(filter_col == filter_val, CompChapterActivity.is_published == True)
    )
    total_questions = _count_result.first()
    if isinstance(total_questions, tuple):
        total_questions = total_questions[0]
    if total_questions == 0:
        raise HTTPException(status_code=404, detail="No activities found")

    play_session = CompActivityPlaySession(
        user_id=current_user.id,
        comp_chapter_id=comp_chapter_id,
        sub_chapter_id=sub_chapter_id,
        total_questions=total_questions,
    )
    session.add(play_session)
    await session.commit()
    await session.refresh(play_session)

    _result = await session.exec(
        select(CompChapterActivity)
        .where(filter_col == filter_val, CompChapterActivity.is_published == True)
        .order_by(*sort_ordering(CompChapterActivity))
        .limit(1)
    )
    first_activity = _result.first()
    return {"data": {"session": play_session.dict(), "next_activity": first_activity.dict() if first_activity else None}}


@router.post("/sessions/{session_id}/answers")
async def submit_comp_answer(
    session_id: int,
    payload: CompAnswerSubmit,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    play_session = await session.get(CompActivityPlaySession, session_id)
    if not play_session:
        raise HTTPException(status_code=404, detail="Session not found")
    if play_session.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not allowed")
    if play_session.status == "completed":
        raise HTTPException(status_code=400, detail="Session already completed")

    activity = await session.get(CompChapterActivity, payload.activity_id)
    if not activity or (activity.comp_chapter_id != play_session.comp_chapter_id and activity.sub_chapter_id != play_session.sub_chapter_id):
        raise HTTPException(status_code=404, detail="Activity not found for this session")
    if not activity.is_published:
        raise HTTPException(status_code=400, detail="Activity not published")

    _existing = await session.exec(
        select(CompActivityAnswer).where(
            CompActivityAnswer.session_id == session_id,
            CompActivityAnswer.activity_id == payload.activity_id,
        )
    )
    if _existing.first():
        raise HTTPException(status_code=400, detail="Answer already submitted")

    if payload.selected_option_index not in {1, 2, 3, 4}:
        raise HTTPException(status_code=400, detail="selected_option_index must be 1-4")

    is_correct = payload.selected_option_index == activity.correct_option_index
    score = 1 if is_correct else 0

    answer = CompActivityAnswer(
        session_id=session_id,
        activity_id=payload.activity_id,
        selected_option_index=payload.selected_option_index,
        is_correct=is_correct,
        score=score,
    )
    session.add(answer)
    play_session.correct_count += 1 if is_correct else 0
    play_session.score += score
    session.add(play_session)
    await session.commit()
    await session.refresh(play_session)

    _answered_count = await session.exec(
        select(func.count(CompActivityAnswer.id)).where(CompActivityAnswer.session_id == session_id)
    )
    answered_count = _answered_count.first()
    if isinstance(answered_count, tuple):
        answered_count = answered_count[0]

    completed = answered_count >= play_session.total_questions
    if completed:
        play_session.status = "completed"
        play_session.completed_at = datetime.now(UTC)
        session.add(play_session)
        await session.commit()
        await session.refresh(play_session)

    _answered_ids = await session.exec(
        select(CompActivityAnswer.activity_id).where(CompActivityAnswer.session_id == session_id)
    )
    answered_ids = [row[0] if isinstance(row, tuple) else row for row in _answered_ids.all()]

    filter_col = CompChapterActivity.comp_chapter_id if play_session.comp_chapter_id else CompChapterActivity.sub_chapter_id
    filter_val = play_session.comp_chapter_id or play_session.sub_chapter_id
    next_query = select(CompChapterActivity).where(filter_col == filter_val, CompChapterActivity.is_published == True)
    if answered_ids:
        next_query = next_query.where(~CompChapterActivity.id.in_(answered_ids))
    next_query = next_query.order_by(*sort_ordering(CompChapterActivity)).limit(1)
    _next = await session.exec(next_query)
    next_activity = _next.first()

    correct_answer = {
        "correct_option_index": activity.correct_option_index,
        "correct_option_text": (
            activity.options[activity.correct_option_index - 1]
            if activity.options and activity.correct_option_index
            else None
        ),
        "answer_description": activity.answer_description,
    }

    return {"data": {
        "is_correct": is_correct,
        "score": score,
        "correct_answer": correct_answer,
        "answer_image_url": activity.answer_image_url,
        "next_activity": next_activity.dict() if next_activity else None,
        "completed": completed,
        "session": play_session.dict(),
    }}


@router.get("/sessions/{session_id}/report")
async def get_comp_session_report(
    session_id: int,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    play_session = await session.get(CompActivityPlaySession, session_id)
    if not play_session:
        raise HTTPException(status_code=404, detail="Session not found")
    if play_session.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not allowed")

    _result = await session.exec(
        select(CompActivityAnswer, CompChapterActivity)
        .join(CompChapterActivity, CompChapterActivity.id == CompActivityAnswer.activity_id)
        .where(CompActivityAnswer.session_id == session_id)
        .order_by(*sort_ordering(CompChapterActivity))
    )
    answers = []
    for answer, activity in _result.all():
        answers.append({
            "activity_id": activity.id,
            "type": activity.type,
            "question_text": activity.question_text,
            "options": activity.options,
            "answer_image_url": activity.answer_image_url,
            "submitted": {"selected_option_index": answer.selected_option_index},
            "correct": {
                "correct_option_index": activity.correct_option_index,
                "correct_option_text": (
                    activity.options[activity.correct_option_index - 1]
                    if activity.options and activity.correct_option_index
                    else None
                ),
                "answer_description": activity.answer_description,
            },
            "is_correct": answer.is_correct,
            "score": answer.score,
        })

    return {"data": {"session": play_session.dict(), "answers": answers}}
