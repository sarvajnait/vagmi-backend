import json
import random
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Any, Dict, List, Optional

from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.messages import HumanMessage, SystemMessage
from langchain_text_splitters import RecursiveCharacterTextSplitter
from loguru import logger
from pydantic import BaseModel, conlist

from app.core.agents.graph import merge_chunks_remove_overlap

BOARD_TEXTBOOK_COLLECTION = "llm_textbooks"
BOARD_QA_COLLECTION = "qa_patterns"
COMP_TEXTBOOK_COLLECTION = "comp_llm_textbooks"
COMP_QA_COLLECTION = "comp_qa_patterns"

MAX_TOPIC_CHARS = 12000
MAX_ACTIVITY_CHUNK_CHARS = 12000
LLM_RETRY_MAX_ATTEMPTS = 3
LLM_RETRY_BASE_DELAY_SEC = 1.5
_RETRYABLE_LLM_ERROR_PATTERNS = [
    re.compile(pattern, re.IGNORECASE)
    for pattern in [
        r"\b429\b",
        r"\b5\d{2}\b",
        r"rate limit",
        r"resource exhausted",
        r"quota",
        r"deadline exceeded",
        r"timed?\s*out",
        r"timeout",
        r"temporar",
        r"unavailable",
        r"connection reset",
        r"connection aborted",
        r"service unavailable",
        r"internal error",
    ]
]


def _extract_json(text: str) -> Dict[str, Any]:
    if not text:
        raise ValueError("Empty AI response")
    match = re.search(r"\{.*\}", text, flags=re.DOTALL)
    if not match:
        raise ValueError("No JSON object found in AI response")
    return json.loads(match.group(0))


def _is_retryable_llm_error(exc: Exception) -> bool:
    message = str(exc)
    return any(pattern.search(message) for pattern in _RETRYABLE_LLM_ERROR_PATTERNS)


def _invoke_llm_with_retry(llm, messages, operation_name: str):
    last_exc: Optional[Exception] = None
    for attempt in range(1, LLM_RETRY_MAX_ATTEMPTS + 1):
        try:
            return llm.invoke(messages)
        except Exception as exc:
            last_exc = exc
            should_retry = _is_retryable_llm_error(exc) and attempt < LLM_RETRY_MAX_ATTEMPTS
            if not should_retry:
                raise
            delay = LLM_RETRY_BASE_DELAY_SEC * (2 ** (attempt - 1)) + random.uniform(0, 0.5)
            logger.warning(
                f"{operation_name} failed with retryable error "
                f"(attempt {attempt}/{LLM_RETRY_MAX_ATTEMPTS}): {exc}. "
                f"Retrying in {delay:.1f}s"
            )
            time.sleep(delay)
    if last_exc:
        raise last_exc
    raise RuntimeError(f"{operation_name} failed without a captured exception")


# --------------------
# LLM helpers
# --------------------

def _get_llm() -> ChatGoogleGenerativeAI:
    return ChatGoogleGenerativeAI(
        model="gemini-2.5-flash",
        temperature=0.2,
        streaming=False,
    )


def _split_text(text: str, chunk_size: int, chunk_overlap: int) -> List[str]:
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size, chunk_overlap=chunk_overlap
    )
    return splitter.split_text(text)


# --------------------
# Context fetchers
# --------------------

def get_full_chapter_text(chapter_id: int, collection_name: str = BOARD_TEXTBOOK_COLLECTION) -> str:
    from app.core.config import settings
    import psycopg

    logger.debug(f"[chapter={chapter_id}] Fetching full chapter text from DB (collection={collection_name!r})")
    try:
        postgres_url = settings.POSTGRES_URL.replace("postgresql+psycopg://", "postgresql://")

        with psycopg.connect(postgres_url) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT document, cmetadata
                    FROM langchain_pg_embedding
                    WHERE collection_id = (
                        SELECT uuid FROM langchain_pg_collection WHERE name = %s
                    )
                    AND cmetadata->>'chapter_id' = %s
                    ORDER BY (cmetadata->>'chunk_index')::int NULLS LAST
                    """,
                    (collection_name, str(chapter_id)),
                )
                rows = cur.fetchall()

                if not rows:
                    logger.warning(f"[chapter={chapter_id}] No embeddings found in {collection_name!r} — chapter has no uploaded content")
                    return ""

                raw_chunks = [row[0] for row in rows]
                text = merge_chunks_remove_overlap(raw_chunks, overlap_chars=200)
                logger.debug(f"[chapter={chapter_id}] Fetched {len(rows)} chunks, merged text length={len(text)}")
                return text
    except Exception as e:
        logger.error(f"[chapter={chapter_id}] Error fetching chapter text from DB: {e} — falling back to similarity search on {collection_name!r}")
        from app.core.config import settings as _settings
        from langchain_postgres import PGVector
        from langchain_google_genai import GoogleGenerativeAIEmbeddings
        fallback_store = PGVector(
            connection=_settings.POSTGRES_URL,
            collection_name=collection_name,
            embeddings=GoogleGenerativeAIEmbeddings(model="models/gemini-embedding-001", output_dimensionality=768),
        )
        docs = fallback_store.similarity_search(
            query="",
            k=1000,
            filter={"chapter_id": str(chapter_id)},
        )
        if not docs:
            logger.warning(f"[chapter={chapter_id}] Fallback similarity search on {collection_name!r} also returned no docs")
            return ""
        docs.sort(key=lambda d: d.metadata.get("chunk_index", 0))
        raw_chunks = [d.page_content for d in docs]
        return merge_chunks_remove_overlap(raw_chunks, overlap_chars=200)



def get_qa_context(chapter_id: int, collection_name: str = BOARD_QA_COLLECTION) -> str:
    """Fetch all Q&A pattern content for a chapter directly from the DB."""
    from app.core.config import settings
    import psycopg

    try:
        postgres_url = settings.POSTGRES_URL.replace("postgresql+psycopg://", "postgresql://")
        with psycopg.connect(postgres_url) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT document
                    FROM langchain_pg_embedding
                    WHERE collection_id = (
                        SELECT uuid FROM langchain_pg_collection WHERE name = %s
                    )
                    AND cmetadata->>'chapter_id' = %s
                    ORDER BY (cmetadata->>'chunk_index')::int NULLS LAST
                    """,
                    (collection_name, str(chapter_id)),
                )
                rows = cur.fetchall()
                if not rows:
                    logger.debug(f"[chapter={chapter_id}] No QA patterns found")
                    return ""
                logger.debug(f"[chapter={chapter_id}] Fetched {len(rows)} QA pattern chunks")
                return "\n\n".join([row[0] for row in rows])
    except Exception as e:
        logger.warning(f"[chapter={chapter_id}] Failed to fetch QA patterns: {e}")
        return ""


# --------------------
# Topic generation
# --------------------

def _language_instruction(medium_name: str, context: str = "topic titles and summaries") -> str:
    medium_lower = medium_name.lower()
    if any(
        lang in medium_lower
        for lang in [
            "english", "hindi", "kannada", "malayalam", "tamil",
            "telugu", "sanskrit", "urdu", "bengali", "marathi", "gujarati",
        ]
    ):
        lang_name = medium_name.replace(" medium", "").replace("Medium", "").strip()
        return (
            f"\nIMPORTANT: This is {lang_name} medium. Generate all {context} "
            f"in {lang_name} language only."
        )
    return ""


def _generate_topics_from_text(
    text: str, medium_name: str = ""
) -> List[Dict[str, str]]:
    human_prompt = (
        "Analyze the provided chapter text and extract a comprehensive yet concise list of key topics "
        "that represent the entire scope of the content.\n\n"
        "Requirements:\n"
        "- Optimized Topic Count: Extract between 10 to 15 topics depending on the chapter's length and complexity. "
        "Prefer more topics over fewer — only merge if two topics are nearly identical.\n"
        "- Complete Coverage: Ensure that every concept, definition, and principle mentioned "
        "in the chapter is mapped to one of the topics. No part of the chapter should be left out.\n"
        "- Conceptual Pillars: Each topic name must be a significant 'Conceptual Pillar' rather than a minor detail.\n"
        "- Exhaustive Logic: The final list must be structured so that 'Important Questions' generated for these "
        "topics will collectively cover the entire chapter without any gaps.\n\n"
        "Keep titles short and summaries 1 sentence.\n"
        'Return JSON in this schema only: { "topics": [ { "title": "...", "summary": "..." } ] }\n'
        f"{_language_instruction(medium_name)}\n\n"
        f"MEDIUM: {medium_name}\n"
        f"CHAPTER TEXT:\n{text}"
    )

    llm = _get_llm()
    response = _invoke_llm_with_retry(
        llm,
        [
            SystemMessage(content="Act as an expert Curriculum Designer. Your output must be strict JSON only ? no markdown, no prose, no extra text."),
            HumanMessage(content=human_prompt),
        ],
        operation_name="Topic generation",
    )
    payload = _extract_json(response.content)
    topics = payload.get("topics", [])
    return topics if isinstance(topics, list) else []


def _consolidate_topics(
    topics: List[Dict[str, str]],
    medium_name: str = "",
) -> List[Dict[str, str]]:
    human_prompt = (
        "Given the topic candidates below extracted from different parts of a chapter, "
        "deduplicate and consolidate them into the final comprehensive topic list.\n\n"
        "Requirements:\n"
        "- Optimized Topic Count: Return between 10 to 15 topics depending on the chapter's breadth. "
        "Prefer more topics over fewer — only merge if two topics are nearly identical.\n"
        "- Complete Coverage: Preserve all distinct concepts — do not drop topics that represent unique content.\n"
        "- Conceptual Pillars: Each topic name must be a significant 'Conceptual Pillar' rather than a minor detail.\n"
        "- Deduplication: Merge topics that refer to the same concept into one well-named topic.\n"
        "- Exhaustive Logic: The final list must be thorough enough that generating 'Important Questions' "
        "for each topic would cover the entire chapter without gaps.\n\n"
        'Return JSON in this schema only: { "topics": [ { "title": "...", "summary": "..." } ] }\n'
        f"{_language_instruction(medium_name)}\n\n"
        f"MEDIUM: {medium_name}\n\n"
        f"TOPICS:\n{json.dumps(topics)}"
    )

    llm = _get_llm()
    response = _invoke_llm_with_retry(
        llm,
        [
            SystemMessage(content="Act as an expert Curriculum Designer. Your output must be strict JSON only ? no markdown, no prose, no extra text."),
            HumanMessage(content=human_prompt),
        ],
        operation_name="Topic consolidation",
    )
    payload = _extract_json(response.content)
    topics = payload.get("topics", [])
    return topics if isinstance(topics, list) else []


def generate_topics(chapter_id: int, medium_name: str = "", collection_name: str = BOARD_TEXTBOOK_COLLECTION) -> List[Dict[str, str]]:
    logger.info(f"[chapter={chapter_id}] Starting topic generation (medium={medium_name!r}, collection={collection_name!r})")
    chapter_text = get_full_chapter_text(chapter_id, collection_name)
    if not chapter_text:
        logger.warning(f"[chapter={chapter_id}] No chapter text — skipping topic generation")
        return []

    logger.info(f"[chapter={chapter_id}] Chapter text length={len(chapter_text)}")
    all_topics: List[Dict[str, str]] = []
    if len(chapter_text) <= MAX_TOPIC_CHARS:
        logger.debug(f"[chapter={chapter_id}] Single chunk topic extraction")
        all_topics.extend(_generate_topics_from_text(chapter_text, medium_name))
        logger.info(f"[chapter={chapter_id}] Topics before consolidation: {len(all_topics)}")
        final = _consolidate_topics(all_topics, medium_name=medium_name)
        logger.info(f"[chapter={chapter_id}] Final topics after consolidation: {len(final)}")
        return final

    chunks = _split_text(chapter_text, chunk_size=MAX_TOPIC_CHARS, chunk_overlap=400)
    logger.info(f"[chapter={chapter_id}] Chapter too large, split into {len(chunks)} chunks for parallel extraction")
    with ThreadPoolExecutor(max_workers=min(len(chunks), 5)) as executor:
        futures = {
            executor.submit(_generate_topics_from_text, chunk, medium_name): i
            for i, chunk in enumerate(chunks)
        }
        chunk_results: dict[int, list] = {}
        for future in as_completed(futures):
            idx = futures[future]
            try:
                chunk_results[idx] = future.result()
                logger.debug(f"[chapter={chapter_id}] Chunk {idx} returned {len(chunk_results[idx])} topics")
            except Exception as exc:
                logger.warning(f"[chapter={chapter_id}] Topic extraction failed for chunk {idx}: {exc}")
                chunk_results[idx] = []
        for i in range(len(chunks)):
            all_topics.extend(chunk_results.get(i, []))

    logger.info(f"[chapter={chapter_id}] Total topics before consolidation: {len(all_topics)}")
    if not all_topics:
        return []
    final = _consolidate_topics(all_topics, medium_name=medium_name)
    logger.info(f"[chapter={chapter_id}] Final topics after consolidation: {len(final)}")
    return final


# --------------------
# Activity generation — Pydantic structured output
# --------------------

class _MCQActivity(BaseModel):
    question_text: str
    options: conlist(str, min_length=4, max_length=4)
    correct_answer: str
    answer_description: str

class _DescriptiveActivity(BaseModel):
    question_text: str
    answer_text: str

class _TopicActivities(BaseModel):
    topic: str
    mcqs: List[_MCQActivity]
    descriptives: List[_DescriptiveActivity]

class _ActivityOutput(BaseModel):
    topics: List[_TopicActivities]


def generate_activities(
    chapter_id: int,
    topic_titles: List[str],
    mcq_count: int,
    descriptive_count: int,
    medium_name: str = "",
    collection_name: str = BOARD_TEXTBOOK_COLLECTION,
    qa_collection_name: str = BOARD_QA_COLLECTION,
) -> List[Dict[str, Any]]:
    """
    Generate activities for all topic titles in a single structured LLM call.
    Returns a flat list of activities, each with a 'topic' field.
    """
    logger.info(
        f"[chapter={chapter_id}] generate_activities: topics={topic_titles}, "
        f"mcq={mcq_count}, descriptive={descriptive_count}, medium={medium_name!r}, collection={collection_name!r}"
    )
    chapter_text = get_full_chapter_text(chapter_id, collection_name)
    if not chapter_text:
        logger.warning(f"[chapter={chapter_id}] Empty chapter text — cannot generate activities")
        return []
    logger.info(f"[chapter={chapter_id}] Chapter text length={len(chapter_text)}")

    qa_context = get_qa_context(chapter_id, qa_collection_name)
    topics_str = "\n".join(f"- {t}" for t in topic_titles)

    language_instruction = ""
    medium_lower = medium_name.lower()
    if any(
        lang in medium_lower
        for lang in [
            "kannada", "malayalam", "tamil", "telugu", "hindi",
            "sanskrit", "urdu", "bengali", "marathi", "gujarati",
        ]
    ):
        lang_name = medium_name.replace(" medium", "").replace("Medium", "").strip()
        language_instruction = (
            f"\n\nCRITICAL: This is {lang_name} medium. "
            f"ALL questions, options, and answers MUST be in {lang_name} language."
        )

    human_prompt = (
        f"Generate exactly {mcq_count} MCQ questions and exactly {descriptive_count} descriptive questions "
        f"for EACH of the following topics.\n\n"
        "Objective: Ensure 100% information coverage. Every fact, definition, and important question "
        "from both the TEXTBOOK_CONTENT and the QA_DOCUMENT must be converted into a question. "
        "No detail should be left untested.\n\n"
        "MCQ Typology — distribute across all four types for comprehensive testing:\n"
        "1. Direct Concept: Tests factual knowledge, definitions, or laws directly from the text.\n"
        "2. Situational/Application: A real-world scenario where the concept must be applied to find the solution.\n"
        "3. Assertion & Reason: Two statements (A and R) where the relationship between the fact and its logic must be verified.\n"
        "4. Negative Selection: Questions that ask 'Which of the following is NOT correct?' to test boundary knowledge.\n\n"
        "Strict Requirements:\n"
        "- Source Hierarchy: If a question exists in the QA_DOCUMENT, it must be prioritized and adapted as the first set of MCQs.\n"
        f"- For each topic: exactly {mcq_count} MCQs and exactly {descriptive_count} descriptive questions — do not skip any topic.\n"
        "- MCQ: provide 4 options, correct_answer must be the EXACT text of one option.\n"
        "- Distractor Quality: Options B, C, and D must be plausible 'trap' answers based on common misconceptions related to the topic.\n"
        "- Verifiable Explanations: Every answer_description must include a brief explanation based on the provided source material.\n"
        "- Descriptive: include a full model answer_text.\n"
        f"{language_instruction}\n\n"
        f"MEDIUM: {medium_name}\n\n"
        f"TOPICS:\n{topics_str}\n\n"
        + (
            f"QA_DOCUMENT (Previous Year / Must-Study — prioritize these first):\n{qa_context}\n\n"
            if qa_context else ""
        )
        + f"TEXTBOOK_CONTENT:\n{chapter_text}"
    )

    logger.info(f"[chapter={chapter_id}] Calling LLM with structured output (prompt_len={len(human_prompt)})")
    llm = _get_llm().with_structured_output(_ActivityOutput)
    try:
        result: _ActivityOutput = _invoke_llm_with_retry(
            llm,
            [
                SystemMessage(content=(
                    "Act as a Senior Assessment Designer and Subject Matter Expert. "
                    "Your output must strictly follow the requested JSON schema ? "
                    "no markdown, no prose, no commentary outside the JSON."
                )),
                HumanMessage(content=human_prompt),
            ],
            operation_name=f"Activity generation for chapter {chapter_id}",
        )
    except Exception as exc:
        logger.error(f"[chapter={chapter_id}] Structured output LLM call failed: {exc}")
        return []

    all_activities: List[Dict[str, Any]] = []
    for topic_block in result.topics:
        for mcq in topic_block.mcqs:
            all_activities.append({
                "type": "mcq",
                "topic": topic_block.topic,
                "question_text": mcq.question_text,
                "options": list(mcq.options),
                "correct_answer": mcq.correct_answer,
                "answer_description": mcq.answer_description,
            })
        for desc in topic_block.descriptives:
            all_activities.append({
                "type": "descriptive",
                "topic": topic_block.topic,
                "question_text": desc.question_text,
                "answer_text": desc.answer_text,
            })

    logger.info(f"[chapter={chapter_id}] Structured output returned {len(all_activities)} total activities across {len(result.topics)} topics")
    return all_activities


# --------------------
# Chapter summary (free text — no structured output)
# --------------------

MAX_SUMMARY_CHARS = 20000


def generate_chapter_summary(chapter_id: int, medium_name: str = "", collection_name: str = BOARD_TEXTBOOK_COLLECTION) -> str:
    """
    Generate a structured chapter summary from the full chapter text.
    Returns a markdown-formatted summary string.
    Called once during textbook processing and stored as a ChapterArtifact.
    """
    chapter_text = get_full_chapter_text(chapter_id, collection_name)
    if not chapter_text:
        return ""

    text = chapter_text[:MAX_SUMMARY_CHARS]

    language_instruction = ""
    medium_lower = medium_name.lower()
    if any(
        lang in medium_lower
        for lang in [
            "hindi", "kannada", "malayalam", "tamil", "telugu",
            "sanskrit", "urdu", "bengali", "marathi", "gujarati",
        ]
    ):
        lang_name = medium_name.replace(" medium", "").replace("Medium", "").strip()
        language_instruction = f"\nIMPORTANT: This is {lang_name} medium. Write the entire summary in {lang_name} language."

    human_prompt = (
        "Task: Analyze the provided TEXTBOOK_TEXT and generate a structured summary covering 100% of the topics, sub-topics, and core concepts.\n\n"
        "Structural Requirements:\n\n"
        "**The \"Big Picture\" (60 Seconds):** A 3-bullet point bird's-eye view of why this chapter matters and its primary objective.\n\n"
        "**Topic-Wise Deep Dive (8-10 Minutes):** Organize the content into the 6-15 \"Conceptual Pillars\" identified from the text. For each pillar:\n"
        "- **Core Logic:** A 2-sentence explanation of the \"How/Why.\"\n"
        "- **Keywords & Definitions:** Bold all technical terms that carry marks in board exams.\n"
        "- **The \"Must-Know\" Fact:** One high-probability data point or principle.\n\n"
        "**Visual & Logic Aids (3 Minutes):**\n"
        "- **Comparison Tables:** If the text compares two things (e.g., Mitosis vs. Meiosis), create a Markdown table.\n"
        "- **Formula/Equation Box:** List all mathematical or chemical shorthand.\n"
        "- **Mnemonic Trigger:** Provide one acronym or memory trick for complex lists.\n\n"
        "**The \"Examiner's Warning\" (1 Minute):** List 3 common mistakes students make in this specific chapter (e.g., swapping units, confusing similar terms).\n\n"
        "Formatting & Tone:\n"
        "- Use Strict Markdown (Headings, Bolding, Bullet points).\n"
        "- Tone should be Urgent, Clear, and Encouraging.\n"
        "- Avoid fluff: If a sentence doesn't help a student answer a question, delete it.\n"
        f"{language_instruction}\n\n"
        f"MEDIUM: {medium_name}\n\n"
        f"TEXTBOOK_TEXT:\n{text}"
    )

    llm = _get_llm()
    response = _invoke_llm_with_retry(
        llm,
        [
            SystemMessage(content="You are an elite Academic Content Architect. Your goal is to transform a full chapter into a High-Density Revision Guide designed for a student to master the entire chapter in 10?15 minutes."),
            HumanMessage(content=human_prompt),
        ],
        operation_name=f"Chapter summary generation for chapter {chapter_id}",
    )
    return response.content.strip()


def generate_one_mark_questions(chapter_id: int, medium_name: str = "", collection_name: str = BOARD_TEXTBOOK_COLLECTION) -> str:
    chapter_text = get_full_chapter_text(chapter_id, collection_name)
    if not chapter_text:
        return ""

    text = chapter_text[:MAX_SUMMARY_CHARS]

    language_instruction = ""
    medium_lower = medium_name.lower()
    if any(
        lang in medium_lower
        for lang in [
            "hindi", "kannada", "malayalam", "tamil", "telugu",
            "sanskrit", "urdu", "bengali", "marathi", "gujarati",
        ]
    ):
        lang_name = medium_name.replace(" medium", "").replace("Medium", "").strip()
        language_instruction = f"\nIMPORTANT: This is {lang_name} medium. Write all questions and answers in {lang_name} language."

    human_prompt = (
        "Task: Act as a Board Exam Paper Setter. Scan the provided TEXTBOOK_TEXT and extract every possible "
        "1-mark question and answer to create an exhaustive objective question bank.\n\n"
        "Extraction Typology — generate questions in all four formats:\n\n"
        "1. **Factual Recall (VSA):** Direct questions requiring a one-word or one-sentence answer (e.g., 'Define Power Sharing').\n\n"
        "2. **Assertion & Reason:** Create a statement (Assertion) and a logical explanation (Reason). "
        "Students must decide if R explains A.\n\n"
        "3. **Identify the Correct/Incorrect Pair:** Extract lists or classifications from the text and create "
        "'Match the following' or 'Pick the odd one out' style 1-mark items.\n\n"
        "4. **Term Spotting:** Provide a definition or a function and ask for the specific scientific or technical term.\n\n"
        "Strict Requirements:\n"
        "- **Granularity:** If a paragraph contains three distinct facts (e.g., a date, a name, and a location), create three separate 1-mark questions.\n"
        "- **Accuracy:** Answers must be verbatim from the textbook to ensure they match the official marking scheme.\n"
        "- **No Duplication:** If a fact is covered in one question type, do not repeat it unless the context changes.\n"
        "- **Coverage Check:** Ensure that even the 'Small Print' (box content, captions of diagrams, and footnotes) is converted into 1-mark questions.\n"
        "- Use Strict Markdown formatting. Bold technical terms.\n"
        f"{language_instruction}\n\n"
        f"MEDIUM: {medium_name}\n\n"
        f"TEXTBOOK_TEXT:\n{text}"
    )

    llm = _get_llm()
    response = _invoke_llm_with_retry(
        llm,
        [
            SystemMessage(content="You are a Senior Board Examiner and Curriculum Expert specializing in 1-mark question banks."),
            HumanMessage(content=human_prompt),
        ],
        operation_name=f"One-mark question generation for chapter {chapter_id}",
    )
    return response.content.strip()


def generate_important_questions(chapter_id: int, medium_name: str = "", collection_name: str = BOARD_TEXTBOOK_COLLECTION) -> str:
    chapter_text = get_full_chapter_text(chapter_id, collection_name)
    if not chapter_text:
        return ""

    text = chapter_text[:MAX_SUMMARY_CHARS]

    language_instruction = ""
    medium_lower = medium_name.lower()
    if any(
        lang in medium_lower
        for lang in [
            "hindi", "kannada", "malayalam", "tamil", "telugu",
            "sanskrit", "urdu", "bengali", "marathi", "gujarati",
        ]
    ):
        lang_name = medium_name.replace(" medium", "").replace("Medium", "").strip()
        language_instruction = f"\nIMPORTANT: This is {lang_name} medium. Write all questions and answers in {lang_name} language."

    human_prompt = (
        "Task: Analyze the TEXTBOOK_TEXT and extract all potential descriptive questions that could appear in an exam, "
        "categorized by their marking weightage (2, 3, and 5 marks).\n\n"
        "**2-Mark Questions (Short Answer):** Focus on 'Distinguish between,' 'Give two reasons,' or 'Define and give an example.' "
        "Provide exactly 2-3 concise bullet points for the answer.\n\n"
        "**3-Mark Questions (Analytical):** Focus on 'Explain the process of,' 'Why is [X] important,' or 'Describe the features of.' "
        "Provide 3-4 detailed bullet points with technical keywords.\n\n"
        "**5-Mark Questions (Long/Comprehensive):** Focus on the major themes of the chapter. Provide a structured response:\n"
        "- **Heading:** Clear title for the answer.\n"
        "- **Intro:** A 1-line opening.\n"
        "- **Body:** 5-6 comprehensive bullet points covering the 'How, Why, and What.'\n"
        "- **Conclusion/Note:** A concluding line or mention of a mandatory diagram/formula.\n\n"
        "Strict Constraints:\n"
        "- **Bullet-Only Answers:** All answers must be in clean bullet points for high readability and last-minute memorization.\n"
        "- **Keyword Emphasis:** Bold the most important technical terms in every bullet point.\n"
        "- **Exhaustive Mapping:** Ensure every sub-topic long enough to form a descriptive question is included.\n"
        "- **No Overlap:** If a concept is covered in a 5-mark question, do not create a separate 2-mark question for the same detail unless it's a specific sub-part.\n"
        f"{language_instruction}\n\n"
        f"MEDIUM: {medium_name}\n\n"
        f"TEXTBOOK_TEXT:\n{text}"
    )

    llm = _get_llm()
    response = _invoke_llm_with_retry(
        llm,
        [
            SystemMessage(content="You are a Senior Board Examiner and Curriculum Expert specializing in descriptive question banks."),
            HumanMessage(content=human_prompt),
        ],
        operation_name=f"Important question generation for chapter {chapter_id}",
    )
    return response.content.strip()


# --------------------
# Activity normalization
# --------------------

def normalize_activity(item: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    activity_type = str(item.get("type", "")).strip().lower()
    question_text = str(item.get("question_text", "")).strip()

    if not activity_type or not question_text:
        return None

    if activity_type == "mcq":
        options = item.get("options", [])
        if not isinstance(options, list) or len(options) != 4:
            return None
        cleaned_options = [str(opt).strip() for opt in options]
        if any(not opt for opt in cleaned_options):
            return None
        correct_answer = str(item.get("correct_answer", "")).strip()
        try:
            correct_option_index = cleaned_options.index(correct_answer) + 1
        except ValueError:
            lower_options = [o.lower() for o in cleaned_options]
            try:
                correct_option_index = lower_options.index(correct_answer.lower()) + 1
            except ValueError:
                return None
        answer_description = str(item.get("answer_description", "")).strip() or None
        return {
            "type": "mcq",
            "question_text": question_text,
            "options": cleaned_options,
            "correct_option_index": int(correct_option_index),
            "answer_text": None,
            "answer_description": answer_description,
        }

    if activity_type == "descriptive":
        answer_text = str(item.get("answer_text", "")).strip()
        if not answer_text:
            return None
        return {
            "type": "descriptive",
            "question_text": question_text,
            "options": None,
            "correct_option_index": None,
            "answer_text": answer_text,
        }

    return None


# --------------------
# Descriptive answer evaluation
# --------------------

def evaluate_descriptive_answer(
    question: str,
    correct_answer: str,
    user_answer: str,
    medium_name: str = "",
) -> Dict[str, Any]:
    """
    Evaluate a descriptive answer using AI.

    Returns:
        {
            "score": int (0-100),
            "feedback": List[str] (3-4 bullet points)
        }
    """
    language_instruction = ""
    medium_lower = medium_name.lower()
    if any(
        lang in medium_lower
        for lang in [
            "kannada", "malayalam", "tamil", "telugu", "hindi",
            "sanskrit", "urdu", "bengali", "marathi", "gujarati",
        ]
    ):
        lang_name = medium_name.replace(" medium", "").replace("Medium", "").strip()
        language_instruction = f"\n\nIMPORTANT: Generate feedback in {lang_name} language since this is {lang_name} medium."

    human_prompt = (
        "Evaluate the student's answer compared to the correct answer. "
        'Return JSON in this schema: { "score": 75, "feedback": ["Good: ...", "Improve: ..."] }\n\n'
        "Requirements:\n"
        "- score: 0-100 based on correctness, completeness, and clarity\n"
        "- feedback: Exactly 3-4 bullet points\n"
        "- Start each feedback point with 'Good:' or 'Improve:'\n"
        "- Be specific and constructive\n"
        f"{language_instruction}\n\n"
        f"MEDIUM: {medium_name}\n"
        f"QUESTION: {question}\n\n"
        f"CORRECT ANSWER:\n{correct_answer}\n\n"
        f"STUDENT'S ANSWER:\n{user_answer}"
    )

    llm = _get_llm()
    response = _invoke_llm_with_retry(
        llm,
        [
            SystemMessage(content="You are an assistant that outputs strict JSON only. Do not include markdown or extra text."),
            HumanMessage(content=human_prompt),
        ],
        operation_name="Descriptive answer evaluation",
    )
    payload = _extract_json(response.content)
    score = max(0, min(100, int(payload.get("score", 0))))
    feedback = payload.get("feedback", [])
    if not isinstance(feedback, list):
        feedback = []
    return {
        "score": score,
        "feedback": feedback,
    }


# ============================================================
# Notes Markdown Conversion
# ============================================================

def _extract_docx_content(content_bytes: bytes) -> tuple[str, dict[str, bytes]]:
    """
    Parse DOCX and return:
    - text: document content with IMAGE:filename placeholders inline
    - images: dict mapping filename → raw bytes
    """
    import zipfile
    import xml.etree.ElementTree as ET
    from io import BytesIO

    ns = {
        "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    }
    rels_ns = "http://schemas.openxmlformats.org/package/2006/relationships"

    with zipfile.ZipFile(BytesIO(content_bytes), "r") as archive:
        xml_data = archive.read("word/document.xml")

        # Parse relationships: rId → media filename
        rId_to_file: dict[str, str] = {}
        try:
            rels_xml = archive.read("word/_rels/document.xml.rels")
            rels_root = ET.fromstring(rels_xml)
            for rel in rels_root.findall(f"{{{rels_ns}}}Relationship"):
                if "image" in rel.get("Type", "").lower():
                    target = rel.get("Target", "")  # e.g. media/image1.png
                    rId_to_file[rel.get("Id", "")] = target.replace("../", "word/").lstrip("/")
        except KeyError:
            pass

        # Extract image bytes
        images: dict[str, bytes] = {}
        for archive_name in archive.namelist():
            if archive_name.startswith("word/media/"):
                images[archive_name] = archive.read(archive_name)

    root = ET.fromstring(xml_data)
    segments: list[str] = []

    for para in root.findall(".//w:p", ns):
        # Check if paragraph contains a drawing/image
        drawing = para.find(".//w:drawing", ns)
        if drawing is not None:
            # Find the rId of the embedded image
            blip = drawing.find(".//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}blip")
            if blip is None:
                blip = drawing.find(".//{http://schemas.openxmlformats.org/drawingml/2006/picture}blip")
            if blip is None:
                # Try direct a:blip
                blip = drawing.find(
                    ".//{http://schemas.openxmlformats.org/drawingml/2006/main}blip"
                )
            if blip is not None:
                r_embed = blip.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
                if r_embed and r_embed in rId_to_file:
                    segments.append(f"IMAGE:{rId_to_file[r_embed]}")
                    continue

        text = "".join(node.text or "" for node in para.findall(".//w:t", ns)).strip()
        if text:
            segments.append(text)

    return "\n\n".join(segments), images


def convert_docx_to_notes_markdown(content_bytes: bytes, note_id: int | None = None) -> str:
    from app.utils.files import upload_bytes_to_do
    import mimetypes

    text_with_placeholders, images = _extract_docx_content(content_bytes)
    if not text_with_placeholders.strip():
        raise ValueError("No text content found in DOCX file")

    # Upload images and replace placeholders with CDN URLs
    uploaded: dict[str, str] = {}
    for archive_path, img_bytes in images.items():
        filename = archive_path.split("/")[-1]
        mime = mimetypes.guess_type(filename)[0] or "image/png"
        do_path = f"comp/notes/images/{note_id or 'misc'}"
        try:
            url = upload_bytes_to_do(img_bytes, filename, do_path, content_type=mime)
            uploaded[archive_path] = url
        except Exception as e:
            logger.warning(f"Could not upload DOCX image {filename}: {e}")

    # Replace IMAGE:path placeholders with markdown image syntax
    lines = text_with_placeholders.split("\n\n")
    resolved_lines = []
    for line in lines:
        if line.startswith("IMAGE:"):
            archive_path = line[6:]
            url = uploaded.get(archive_path)
            if url:
                filename = archive_path.split("/")[-1]
                resolved_lines.append(f"![{filename}]({url})")
            # silently drop images that failed to upload
        else:
            resolved_lines.append(line)
    text = "\n\n".join(resolved_lines)

    llm = ChatGoogleGenerativeAI(model="gemini-2.5-flash", temperature=0.1)
    prompt = f"""You are converting educational content for a competitive exam preparation app into Extended Markdown format.

Convert the following document text into clean Extended Markdown.

BLOCK TYPES — use these where appropriate:
- :::formula ... ::: → formulas, equations, key mathematical/scientific rules
- :::shortcut ... ::: → shortcuts, tricks, mnemonics, quick tips
- :::pyq_alert ... ::: → previous year question info, exam frequency notes
- :::mistake ... ::: → common mistakes, errors to avoid
- :::exam_insight ... ::: → exam pattern info, how the topic is tested
- :::solved_example ... ::: → worked examples, solved problems

RULES:
1. Use # for main section headings, ## for sub-sections, ### for sub-sub-sections
2. Use standard Markdown tables (| col | col |) for data tables
3. Use **bold** for key terms, *italic* for emphasis
4. Preserve bullet lists (- item) and numbered lists (1. item)
5. Each block opens with :::type on its own line and closes with ::: on its own line
6. Do NOT wrap normal paragraphs in blocks
7. Keep the content faithful to the original — do not add or remove information
8. ![image](url) tags are already placed at the correct positions — keep them exactly where they are, each on its own line

DOCUMENT TEXT:
{text}

Return ONLY the Extended Markdown. No preamble, no commentary, no code fences."""

    response = _invoke_llm_with_retry(llm, [HumanMessage(content=prompt)], "convert_docx_to_notes_markdown")
    return response.content.strip()


def convert_excel_to_notes_markdown(content_bytes: bytes) -> str:
    from openpyxl import load_workbook
    from io import BytesIO

    wb = load_workbook(filename=BytesIO(content_bytes), data_only=True)
    ws = wb.active
    lines = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        row_type = str(row[0]).strip().lower()
        content = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if not content:
            continue

        if row_type == "heading1":
            lines.append(f"# {content}\n")
        elif row_type == "heading2":
            lines.append(f"## {content}\n")
        elif row_type == "heading3":
            lines.append(f"### {content}\n")
        elif row_type == "paragraph":
            lines.append(f"{content}\n")
        elif row_type == "divider":
            lines.append("---\n")
        elif row_type in ("formula", "shortcut", "pyq_alert", "mistake", "exam_insight", "solved_example"):
            lines.append(f":::{row_type}")
            lines.append(content)
            for extra in row[2:]:
                if extra is not None:
                    lines.append(str(extra).strip())
            lines.append(":::\n")
        elif row_type == "mcq":
            opts = [str(row[i]).strip() if len(row) > i and row[i] is not None else "" for i in range(2, 6)]
            correct = str(row[6]).strip() if len(row) > 6 and row[6] is not None else "0"
            explanation = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""
            lines.append(":::mcq")
            lines.append(f"question: {content}")
            lines.append(f"options: {json.dumps(opts)}")
            lines.append(f"correct: {correct}")
            if explanation:
                lines.append(f"explanation: {explanation}")
            lines.append(":::\n")

    if not lines:
        raise ValueError("No content found in Excel file. Ensure data starts from row 2 with type in column A and content in column B.")

    return "\n".join(lines)
